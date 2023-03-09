from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.feature_extraction.text import CountVectorizer
import pandas as pd
import numpy as np
import openpyxl as op

def tfidf(corpus):
  #vect = TfidfVectorizer()
  vect = CountVectorizer()
  document_term_matrix = vect.fit_transform(corpus)

  #tf = pd.DataFrame(document_term_matrix.toarray(), columns=vect.get_feature_names_out())
  tf = pd.DataFrame(document_term_matrix.toarray(), columns=vect.get_feature_names_out())
                                               # TF (Term Frequency)
  D = len(tf)
  df = tf.astype(bool).sum(axis=0)
  idf = np.log((D+1) / (df+1)) + 1             # IDF (Inverse Document Frequency)

  print(tf)
  print(idf)

  # TF-IDF (Term Frequency-Inverse Document Frequency)
  tfidf = tf * idf
  tfidf = tfidf / np.linalg.norm(tfidf, axis=1, keepdims=True)

  print(tfidf)

def add_value(exist_dict, key, added_values):
  if key not in exist_dict:
    exist_dict[key] = list()

  exist_dict[key].append(added_values)

  return exist_dict

data = {}
char_list = []

def preprocessing(data):
  global char_list

  char_list = list(data.values())
  for i in range(len(data)):
    char_list[i] = ' '.join(char_list[i])

def read_data():
  wb = op.load_workbook("/Users/1110165/duke/dataset/CE1000098326_perfect_char.xlsx")
  ws = wb["original"]

  global data
  for i in range(100): #ws.max_row+1):
    column = 1
    cur = ws.cell(row=i + 2, column=1).value
    add_value(data, cur, ws.cell(row=i + 2, column=5).value)

  preprocessing(data)

read_data()
print(data)
#tfidf(char_list)
#tfidf(docs)